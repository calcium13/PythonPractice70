#coding:utf-8
'''
Created on 2018年9月11日

@author: zh
'''
from openpyxl import load_workbook # 默认可读写，若有需要可以指定write_only和read_only为True 


def contain(target,element):
    for i in range(len(target)):
        if(element[0]==target[i][0] and element[1]==target[i][1]):
            return True
    return False
def contain_IP(target,element):
    for i in range(len(target)):
        if(element[1]==target[i][1]):
            return True
    return False

wb = load_workbook('D:\\2018联研院\\日志分析\\20180515&0714.xlsx') #默认打开的文件为可读写，若有需要可以指定参数read_only为True。

# 获得所有sheet的名称 
print(wb.sheetnames) # 根据sheet名字获得sheet 
sheet = wb['succeed0714'] # 获得sheet名 
print(sheet.title) # 获得当前正在显示的sheet, 也可以用wb.get_active_sheet() sheet = wb.active
b4_too = sheet.cell(row=3, column=2) 
print(b4_too.value)
print(sheet.max_row)
succeed_repeat=[]
for i in range (1,sheet.max_row):
    tp=(sheet.cell(row=i, column=3).value,sheet.cell(row=i,column=4).value)
    #print(tp)
    if(contain(succeed_repeat,tp)):
        #print(tp)
        pass
    else:
        succeed_repeat.append(tp)
print(len(succeed_repeat))
for i in range(len(succeed_repeat)):
    print(succeed_repeat[i])
file=open('D:\\2018联研院\\日志分析\\temp0714.txt','w')
for i in range(len(succeed_repeat)):
    for j in range(len(succeed_repeat)):
        if(succeed_repeat[i][1]==succeed_repeat[j][1] and succeed_repeat[i][0]!=succeed_repeat[j][0]):
            file.write(succeed_repeat[j][1]+','+succeed_repeat[j][0])
            file.write('\n')
        else:
            pass
    #file.write(succeed_repeat[i][1]+','+succeed_repeat[i][0])    
file.flush()
file.close()
